import os
import uuid
import base64
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, jsonify, make_response
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_prefix=1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
SIGNATURES_DIR = os.path.join(DATA_DIR, 'signatures')
PHOTOS_DIR = os.path.join(DATA_DIR, 'photos')
DB_PATH = os.path.join(DATA_DIR, 'mannequins.db')
SECRET_KEY_FILE = os.path.join(DATA_DIR, '.secret_key')

os.makedirs(SIGNATURES_DIR, exist_ok=True)
os.makedirs(PHOTOS_DIR, exist_ok=True)

ALLOWED_PHOTO_EXT = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}

# Persistent secret key
if os.path.exists(SECRET_KEY_FILE):
    with open(SECRET_KEY_FILE, 'rb') as f:
        app.secret_key = f.read()
else:
    app.secret_key = os.urandom(32)
    with open(SECRET_KEY_FILE, 'wb') as f:
        f.write(app.secret_key)

# Durée de la session « se souvenir de moi »
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)

MANNEQUIN_TYPES = ['Homme', 'Femme', 'Enfant', 'Nourrisson']
DEFAULT_ADMIN_PASSWORD = 'ALUPSAdmin'

# Intervenants pré-enregistrés : amorçage initial de la table `intervenants`.
# Format (prénom, nom) — le nom peut comporter plusieurs mots.
DEFAULT_INTERVENANTS = [
    ('Arthur', 'PRUVOST RIVIÈRE'),
    ('Dina', 'FERNANDES'),
    ('Jérôme', 'RIVIÈRE'),
    ('Florence', 'MALAPLATE'),
    ('Thomas', 'GASTELLU'),
    ('Bastien', 'PELLECER'),
    ('Véronique', 'GENOUILLE MONTAGNAC'),
    ('Regnault', 'QUENTIN'),
]

# Statuts possibles d'un mannequin (clé stockée en base -> affichage)
STATUTS = {
    'operationnel': {'label': 'Opérationnel',        'icon': 'circle-check-filled',  'classe': 'operationnel'},
    'hors_service': {'label': 'Hors-Service',         'icon': 'circle-x-filled',      'classe': 'hs'},
    'maintenance':  {'label': 'Maintenance en cours', 'icon': 'tool',                 'classe': 'maintenance'},
    'defaut':       {'label': 'Défaut signalé',       'icon': 'alert-triangle-filled', 'classe': 'defaut'},
}
DEFAULT_STATUT = 'operationnel'

# Types de formation possibles pour une intervention (clé stockée -> libellé)
TYPES_FORMATION = {
    'psc':       'PSC (formation de base)',
    'revisions': 'Révisions',
    'pse':       'PSE',
    'autre':     'Autre',
}
DEFAULT_TYPE_FORMATION = 'psc'


def formation_label(row):
    """Libellé lisible du type de formation d'une intervention (ligne DB).
    Pour « Autre », ajoute la précision saisie si elle existe."""
    keys = row.keys() if hasattr(row, 'keys') else row
    key = (row['type_formation'] if 'type_formation' in keys else '') or ''
    if key == 'autre':
        precision = (row['type_formation_autre'] if 'type_formation_autre' in keys else '') or ''
        return f'Autre — {precision}' if precision else 'Autre'
    return TYPES_FORMATION.get(key, '')


# Rendre le libellé de formation disponible dans les templates
app.jinja_env.globals['formation_label'] = formation_label


# ── Database ──────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS admin (
            id INTEGER PRIMARY KEY,
            password_hash TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS mannequins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('Homme', 'Femme', 'Enfant', 'Nourrisson')),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(numero, type)
        );
        CREATE TABLE IF NOT EXISTS interventions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mannequin_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            prenom TEXT NOT NULL,
            nom TEXT NOT NULL,
            nettoyage INTEGER NOT NULL,
            changement_poumons INTEGER NOT NULL,
            reparation INTEGER NOT NULL,
            informations TEXT DEFAULT '',
            signature_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (mannequin_id) REFERENCES mannequins(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            intervention_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            original_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (intervention_id) REFERENCES interventions(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS intervenants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            prenom TEXT NOT NULL,
            nom TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(prenom, nom)
        );
    ''')
    # Add column if missing (preserves existing data)
    try:
        conn.execute('ALTER TABLE interventions ADD COLUMN description_reparation TEXT DEFAULT ""')
    except sqlite3.OperationalError:
        pass  # Column already exists

    # Type de formation de l'intervention (migration additive, préserve l'existant)
    for ddl in (
        "ALTER TABLE interventions ADD COLUMN type_formation TEXT DEFAULT 'psc'",
        "ALTER TABLE interventions ADD COLUMN type_formation_autre TEXT DEFAULT ''",
    ):
        try:
            conn.execute(ddl)
        except sqlite3.OperationalError:
            pass  # Column already exists

    # Statut des mannequins (migration additive, préserve l'existant)
    for ddl in (
        "ALTER TABLE mannequins ADD COLUMN statut TEXT DEFAULT 'operationnel'",
        "ALTER TABLE mannequins ADD COLUMN statut_detail TEXT DEFAULT ''",
        "ALTER TABLE mannequins ADD COLUMN statut_updated_at TIMESTAMP",
    ):
        try:
            conn.execute(ddl)
        except sqlite3.OperationalError:
            pass  # Column already exists

    # Migration des types : 'Adulte' -> 'Homme' + ajout de 'Femme'.
    # La contrainte CHECK est figée dans la table : on recrée donc la table.
    # foreign_keys OFF pour éviter que le DROP TABLE ne supprime en cascade les
    # interventions ; les id sont conservés pour garder les liens intacts.
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='mannequins'"
    ).fetchone()
    if row and 'Adulte' in row[0]:
        conn.commit()
        conn.execute('PRAGMA foreign_keys = OFF')
        conn.executescript('''
            DROP TABLE IF EXISTS mannequins_new;
            CREATE TABLE mannequins_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero TEXT NOT NULL,
                type TEXT NOT NULL CHECK(type IN ('Homme', 'Femme', 'Enfant', 'Nourrisson')),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                statut TEXT DEFAULT 'operationnel',
                statut_detail TEXT DEFAULT '',
                statut_updated_at TIMESTAMP,
                UNIQUE(numero, type)
            );
            INSERT INTO mannequins_new (id, numero, type, created_at, statut, statut_detail, statut_updated_at)
                SELECT id, numero,
                       CASE WHEN type = 'Adulte' THEN 'Homme' ELSE type END,
                       created_at, statut, statut_detail, statut_updated_at
                FROM mannequins;
            DROP TABLE mannequins;
            ALTER TABLE mannequins_new RENAME TO mannequins;
        ''')
        conn.execute('PRAGMA foreign_keys = ON')

    # Amorçage de la liste d'intervenants — uniquement si la table est vide,
    # afin de ne pas réintroduire ceux supprimés ensuite par l'administrateur.
    if conn.execute('SELECT COUNT(*) FROM intervenants').fetchone()[0] == 0:
        conn.executemany(
            'INSERT OR IGNORE INTO intervenants (prenom, nom) VALUES (?, ?)',
            DEFAULT_INTERVENANTS
        )

    # Correction du nom « RIVIERE » -> « RIVIÈRE » sur les tables déjà amorcées
    # (idempotent : sans effet si le nom est déjà correct ou absent).
    for old, new in (('RIVIERE', 'RIVIÈRE'), ('PRUVOST RIVIERE', 'PRUVOST RIVIÈRE')):
        try:
            conn.execute('UPDATE intervenants SET nom = ? WHERE nom = ?', (new, old))
        except sqlite3.IntegrityError:
            pass  # un intervenant « RIVIÈRE » existe déjà, on laisse tel quel

    conn.commit()
    conn.close()


# ── Auth ──────────────────────────────────────────────────

def is_admin_setup():
    conn = get_db()
    admin = conn.execute('SELECT id FROM admin LIMIT 1').fetchone()
    conn.close()
    return admin is not None


def get_intervenants_ordered():
    """Intervenants avec les plus actifs en tête : tri par nombre
    d'interventions décroissant, puis alphabétique. Le décompte sert
    uniquement au classement et n'est pas renvoyé (aucun indicateur affiché)."""
    conn = get_db()
    rows = conn.execute('''
        SELECT iv.id, iv.prenom, iv.nom
        FROM intervenants iv
        LEFT JOIN interventions i
               ON i.prenom = iv.prenom AND i.nom = iv.nom
        GROUP BY iv.id, iv.prenom, iv.nom
        ORDER BY COUNT(i.id) DESC,
                 iv.nom COLLATE NOCASE, iv.prenom COLLATE NOCASE
    ''').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


# ── Public routes ─────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('formulaire'))


@app.route('/formulaire', methods=['GET'])
def formulaire():
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM mannequins ORDER BY type, numero'
    ).fetchall()
    conn.close()
    mannequins = [dict(r) for r in rows]

    # URL params for direct link
    pre_type = request.args.get('type', '')
    pre_numero = request.args.get('numero', '')

    # Find matching mannequin if params provided
    pre_mannequin_id = None
    if pre_type and pre_numero:
        for m in mannequins:
            if m['type'] == pre_type and m['numero'] == pre_numero:
                pre_mannequin_id = m['id']
                break

    return render_template(
        'form.html',
        mannequins=mannequins,
        intervenants=get_intervenants_ordered(),
        types=MANNEQUIN_TYPES,
        types_formation=TYPES_FORMATION,
        default_type_formation=DEFAULT_TYPE_FORMATION,
        pre_type=pre_type,
        pre_numero=pre_numero,
        pre_mannequin_id=pre_mannequin_id,
        today=datetime.now().strftime('%Y-%m-%d')
    )


@app.route('/formulaire', methods=['POST'])
def formulaire_submit():
    mannequin_id = request.form.get('mannequin_id')
    date = request.form.get('date')
    intervenant_id = request.form.get('intervenant_id', '')
    prenom = request.form.get('prenom', '').strip()
    nom = request.form.get('nom', '').strip()

    # Intervenant : soit choisi dans la liste prédéfinie, soit saisi à la main.
    # Un intervenant sélectionné prime sur les champs texte.
    if intervenant_id.isdigit():
        conn = get_db()
        iv = conn.execute(
            'SELECT prenom, nom FROM intervenants WHERE id = ?', (intervenant_id,)
        ).fetchone()
        conn.close()
        if iv:
            prenom, nom = iv['prenom'], iv['nom']

    type_formation = request.form.get('type_formation', '')
    type_formation_autre = request.form.get('type_formation_autre', '').strip()
    nettoyage = request.form.get('nettoyage')
    changement_poumons = request.form.get('changement_poumons')
    reparation = request.form.get('reparation')
    description_reparation = request.form.get('description_reparation', '').strip()
    informations = request.form.get('informations', '').strip()
    signature_data = request.form.get('signature', '')

    # Validation
    errors = []
    if not mannequin_id:
        errors.append('Veuillez sélectionner un mannequin.')
    if not date:
        errors.append('Veuillez indiquer la date.')
    if not prenom:
        errors.append('Veuillez indiquer votre prénom.')
    if not nom:
        errors.append('Veuillez indiquer votre nom.')
    if type_formation not in TYPES_FORMATION:
        errors.append('Veuillez indiquer le type de formation.')
    if type_formation == 'autre' and not type_formation_autre:
        errors.append('Veuillez préciser le type de formation.')
    if nettoyage not in ('oui', 'non'):
        errors.append('Veuillez indiquer si un nettoyage a été effectué.')
    if changement_poumons not in ('oui', 'non'):
        errors.append('Veuillez indiquer si un changement des poumons a été effectué.')
    if reparation not in ('oui', 'non'):
        errors.append('Veuillez indiquer si une réparation a été effectuée.')
    if reparation == 'oui' and not description_reparation:
        errors.append('Veuillez décrire la réparation effectuée.')
    if not signature_data:
        errors.append('Veuillez signer le formulaire.')

    if errors:
        conn = get_db()
        mannequins = [dict(r) for r in conn.execute(
            'SELECT * FROM mannequins ORDER BY type, numero'
        ).fetchall()]
        conn.close()
        for e in errors:
            flash(e, 'danger')
        return render_template(
            'form.html',
            mannequins=mannequins,
            intervenants=get_intervenants_ordered(),
            types=MANNEQUIN_TYPES,
            types_formation=TYPES_FORMATION,
            default_type_formation=DEFAULT_TYPE_FORMATION,
            pre_type=request.form.get('type_select', ''),
            pre_numero=request.form.get('numero_select', ''),
            pre_mannequin_id=mannequin_id,
            today=date or datetime.now().strftime('%Y-%m-%d'),
            form_data=request.form
        )

    # Save signature
    signature_path = None
    if signature_data and signature_data.startswith('data:image'):
        header, encoded = signature_data.split(',', 1)
        img_bytes = base64.b64decode(encoded)
        filename = f"{uuid.uuid4().hex}.png"
        signature_path = os.path.join(SIGNATURES_DIR, filename)
        with open(signature_path, 'wb') as f:
            f.write(img_bytes)
        signature_path = filename  # Store only filename

    # Save to DB
    conn = get_db()
    cursor = conn.execute(
        '''INSERT INTO interventions
           (mannequin_id, date, prenom, nom, type_formation, type_formation_autre,
            nettoyage, changement_poumons,
            reparation, description_reparation, informations, signature_path)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (
            mannequin_id, date, prenom, nom,
            type_formation,
            type_formation_autre if type_formation == 'autre' else '',
            1 if nettoyage == 'oui' else 0,
            1 if changement_poumons == 'oui' else 0,
            1 if reparation == 'oui' else 0,
            description_reparation if reparation == 'oui' else '',
            informations, signature_path
        )
    )
    intervention_id = cursor.lastrowid

    # Save photos
    photos = request.files.getlist('photos')
    for photo in photos:
        if photo and photo.filename:
            ext = os.path.splitext(photo.filename)[1].lower()
            if ext in ALLOWED_PHOTO_EXT:
                fname = f"{uuid.uuid4().hex}{ext}"
                photo.save(os.path.join(PHOTOS_DIR, fname))
                conn.execute(
                    'INSERT INTO photos (intervention_id, filename, original_name) VALUES (?, ?, ?)',
                    (intervention_id, fname, photo.filename)
                )

    conn.commit()
    conn.close()

    return redirect(url_for('formulaire_success'))


@app.route('/formulaire/success')
def formulaire_success():
    return render_template('form_success.html')


@app.route('/etat')
def etat():
    """Dashboard public de l'état des mannequins (lecture seule, sans auth)."""
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM mannequins ORDER BY type, numero'
    ).fetchall()
    conn.close()

    # Regroupement par type, en conservant l'ordre de MANNEQUIN_TYPES
    mannequins_par_type = {t: [] for t in MANNEQUIN_TYPES}
    for r in rows:
        m = dict(r)
        if m['type'] in mannequins_par_type:
            mannequins_par_type[m['type']].append(m)

    return render_template(
        'etat.html',
        mannequins_par_type=mannequins_par_type,
        types=MANNEQUIN_TYPES,
        statuts=STATUTS,
        default_statut=DEFAULT_STATUT,
        total=len(rows)
    )


# ── PWA (manifest, service worker, offline) ───────────────

@app.route('/manifest.webmanifest')
def manifest():
    resp = make_response(render_template('manifest.webmanifest'))
    resp.headers['Content-Type'] = 'application/manifest+json'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


@app.route('/sw.js')
def service_worker():
    # Servi via Flask (et non en statique) pour que url_for respecte le
    # sous-chemin de déploiement (/ALUPS-Mannequin/) : le scope du service
    # worker devient alors automatiquement ce sous-chemin.
    resp = make_response(render_template('sw.js'))
    resp.headers['Content-Type'] = 'application/javascript'
    resp.headers['Service-Worker-Allowed'] = url_for('index')
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


@app.route('/offline')
def offline():
    return render_template('offline.html')


# ── API ───────────────────────────────────────────────────

@app.route('/api/mannequins')
def api_mannequins():
    type_filter = request.args.get('type', '')
    conn = get_db()
    if type_filter:
        mannequins = conn.execute(
            'SELECT id, numero, type FROM mannequins WHERE type = ? ORDER BY numero',
            (type_filter,)
        ).fetchall()
    else:
        mannequins = conn.execute(
            'SELECT id, numero, type FROM mannequins ORDER BY type, numero'
        ).fetchall()
    conn.close()
    return jsonify([dict(m) for m in mannequins])


# ── Admin routes ──────────────────────────────────────────

@app.route('/admin')
def admin_index():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/setup', methods=['GET', 'POST'])
def admin_setup():
    if is_admin_setup():
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')

        if len(password) < 4:
            flash('Le mot de passe doit contenir au moins 4 caractères.', 'danger')
            return render_template('admin_setup.html')

        if password != confirm:
            flash('Les mots de passe ne correspondent pas.', 'danger')
            return render_template('admin_setup.html')

        conn = get_db()
        conn.execute(
            'INSERT INTO admin (password_hash) VALUES (?)',
            (generate_password_hash(password),)
        )
        conn.commit()
        conn.close()

        session['admin_logged_in'] = True
        flash('Compte administrateur créé avec succès.', 'success')
        return redirect(url_for('admin_dashboard'))

    return render_template('admin_setup.html')


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password', '')
        conn = get_db()
        admin = conn.execute('SELECT password_hash FROM admin LIMIT 1').fetchone()
        conn.close()

        # Check stored password or default fallback
        valid = False
        if admin and check_password_hash(admin['password_hash'], password):
            valid = True
        elif password == DEFAULT_ADMIN_PASSWORD:
            valid = True

        if valid:
            session['admin_logged_in'] = True
            # Se souvenir de moi : session persistante (30 j) sinon expire à la fermeture
            session.permanent = bool(request.form.get('remember'))
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Mot de passe incorrect.', 'danger')

    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))


@app.route('/admin/password', methods=['GET', 'POST'])
@admin_required
def admin_password():
    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')

        if len(password) < 4:
            flash('Le mot de passe doit contenir au moins 4 caractères.', 'danger')
            return render_template('admin_password.html')

        if password != confirm:
            flash('Les mots de passe ne correspondent pas.', 'danger')
            return render_template('admin_password.html')

        conn = get_db()
        admin = conn.execute('SELECT id FROM admin LIMIT 1').fetchone()
        if admin:
            conn.execute(
                'UPDATE admin SET password_hash = ? WHERE id = ?',
                (generate_password_hash(password), admin['id'])
            )
        else:
            conn.execute(
                'INSERT INTO admin (password_hash) VALUES (?)',
                (generate_password_hash(password),)
            )
        conn.commit()
        conn.close()

        flash('Mot de passe modifié avec succès.', 'success')
        return redirect(url_for('admin_dashboard'))

    return render_template('admin_password.html')


@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    conn = get_db()
    mannequins = conn.execute('''
        SELECT m.*, COUNT(i.id) as nb_interventions
        FROM mannequins m
        LEFT JOIN interventions i ON m.id = i.mannequin_id
        GROUP BY m.id
        ORDER BY m.type, m.numero
    ''').fetchall()

    recent = conn.execute('''
        SELECT i.*, m.numero as m_numero, m.type as m_type
        FROM interventions i
        JOIN mannequins m ON i.mannequin_id = m.id
        ORDER BY i.created_at DESC
        LIMIT 20
    ''').fetchall()
    conn.close()

    export_full, export_by_mannequin = _export_sizes()

    return render_template(
        'admin_dashboard.html',
        mannequins=mannequins,
        intervenants=get_intervenants_ordered(),
        recent=recent,
        types=MANNEQUIN_TYPES,
        statuts=STATUTS,
        default_statut=DEFAULT_STATUT,
        export_full=export_full,
        export_by_mannequin=export_by_mannequin,
        export_default=_size_pair(0, 0, 0)
    )


@app.route('/admin/mannequins/add', methods=['POST'])
@admin_required
def admin_add_mannequin():
    m_type = request.form.get('type', '')
    numero = request.form.get('numero', '').strip()

    if m_type not in MANNEQUIN_TYPES:
        flash('Type de mannequin invalide.', 'danger')
        return redirect(url_for('admin_dashboard'))

    if not numero:
        flash('Veuillez indiquer un numéro.', 'danger')
        return redirect(url_for('admin_dashboard'))

    conn = get_db()
    try:
        conn.execute(
            'INSERT INTO mannequins (numero, type) VALUES (?, ?)',
            (numero, m_type)
        )
        conn.commit()
        flash(f'Mannequin {m_type} N°{numero} ajouté.', 'success')
    except sqlite3.IntegrityError:
        flash(f'Le mannequin {m_type} N°{numero} existe déjà.', 'danger')
    finally:
        conn.close()

    return redirect(url_for('admin_dashboard'))


@app.route('/admin/mannequins/<int:mannequin_id>/statut', methods=['POST'])
@admin_required
def admin_update_statut(mannequin_id):
    statut = request.form.get('statut', '')
    detail = request.form.get('statut_detail', '').strip()

    if statut not in STATUTS:
        flash('Statut invalide.', 'danger')
        return redirect(url_for('admin_dashboard'))

    if statut == 'defaut' and not detail:
        flash('Veuillez décrire le défaut (complément d\'information obligatoire).', 'danger')
        return redirect(url_for('admin_dashboard'))

    conn = get_db()
    mannequin = conn.execute(
        'SELECT id FROM mannequins WHERE id = ?', (mannequin_id,)
    ).fetchone()
    if not mannequin:
        conn.close()
        flash('Mannequin introuvable.', 'danger')
        return redirect(url_for('admin_dashboard'))

    conn.execute(
        '''UPDATE mannequins
           SET statut = ?, statut_detail = ?, statut_updated_at = CURRENT_TIMESTAMP
           WHERE id = ?''',
        (statut, detail, mannequin_id)
    )
    conn.commit()
    conn.close()
    flash(f'Statut mis à jour : {STATUTS[statut]["label"]}.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/mannequins/<int:mannequin_id>/delete', methods=['POST'])
@admin_required
def admin_delete_mannequin(mannequin_id):
    conn = get_db()
    # Delete associated signature files and photos
    interventions = conn.execute(
        'SELECT id, signature_path FROM interventions WHERE mannequin_id = ?',
        (mannequin_id,)
    ).fetchall()
    for inter in interventions:
        if inter['signature_path']:
            path = os.path.join(SIGNATURES_DIR, inter['signature_path'])
            if os.path.exists(path):
                os.remove(path)
        inter_photos = conn.execute(
            'SELECT filename FROM photos WHERE intervention_id = ?', (inter['id'],)
        ).fetchall()
        for p in inter_photos:
            path = os.path.join(PHOTOS_DIR, p['filename'])
            if os.path.exists(path):
                os.remove(path)

    conn.execute('DELETE FROM mannequins WHERE id = ?', (mannequin_id,))
    conn.commit()
    conn.close()
    flash('Mannequin supprimé.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/intervenants/add', methods=['POST'])
@admin_required
def admin_add_intervenant():
    prenom = request.form.get('prenom', '').strip()
    nom = request.form.get('nom', '').strip()

    if not prenom or not nom:
        flash('Veuillez indiquer le prénom et le nom de l\'intervenant.', 'danger')
        return redirect(url_for('admin_dashboard'))

    conn = get_db()
    try:
        conn.execute(
            'INSERT INTO intervenants (prenom, nom) VALUES (?, ?)',
            (prenom, nom)
        )
        conn.commit()
        flash(f'Intervenant {prenom} {nom} ajouté.', 'success')
    except sqlite3.IntegrityError:
        flash(f'L\'intervenant {prenom} {nom} existe déjà.', 'danger')
    finally:
        conn.close()

    return redirect(url_for('admin_dashboard'))


@app.route('/admin/intervenants/<int:intervenant_id>/delete', methods=['POST'])
@admin_required
def admin_delete_intervenant(intervenant_id):
    # Retire l'intervenant de la liste prédéfinie. Les interventions déjà
    # enregistrées conservent le prénom/nom saisis : l'historique reste intact.
    conn = get_db()
    conn.execute('DELETE FROM intervenants WHERE id = ?', (intervenant_id,))
    conn.commit()
    conn.close()
    flash('Intervenant retiré de la liste.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/mannequins/<int:mannequin_id>/history')
@admin_required
def admin_history(mannequin_id):
    conn = get_db()
    mannequin = conn.execute(
        'SELECT * FROM mannequins WHERE id = ?', (mannequin_id,)
    ).fetchone()

    if not mannequin:
        flash('Mannequin introuvable.', 'danger')
        conn.close()
        return redirect(url_for('admin_dashboard'))

    interventions = conn.execute('''
        SELECT * FROM interventions
        WHERE mannequin_id = ?
        ORDER BY date DESC, created_at DESC
    ''', (mannequin_id,)).fetchall()

    # Load photos grouped by intervention
    intervention_ids = [i['id'] for i in interventions]
    photos_by_intervention = {}
    if intervention_ids:
        placeholders = ','.join('?' * len(intervention_ids))
        photos = conn.execute(
            f'SELECT * FROM photos WHERE intervention_id IN ({placeholders}) ORDER BY id',
            intervention_ids
        ).fetchall()
        for p in photos:
            iid = p['intervention_id']
            if iid not in photos_by_intervention:
                photos_by_intervention[iid] = []
            photos_by_intervention[iid].append(dict(p))

    conn.close()

    # Tailles estimées de l'export de ce mannequin (sans / avec photos)
    sig_bytes = 0
    for it in interventions:
        if it['signature_path']:
            p = os.path.join(SIGNATURES_DIR, it['signature_path'])
            if os.path.exists(p):
                sig_bytes += os.path.getsize(p)
    photo_bytes = 0
    for plist in photos_by_intervention.values():
        for ph in plist:
            p = os.path.join(PHOTOS_DIR, ph['filename'])
            if os.path.exists(p):
                photo_bytes += os.path.getsize(p)
    export_pair = _size_pair(len(interventions), sig_bytes, photo_bytes)

    return render_template(
        'admin_history.html',
        mannequin=mannequin,
        interventions=interventions,
        photos_by_intervention=photos_by_intervention,
        export_pair=export_pair
    )


@app.route('/admin/interventions/<int:intervention_id>')
@admin_required
def admin_intervention_detail(intervention_id):
    conn = get_db()
    inter = conn.execute('''
        SELECT i.*, m.numero as m_numero, m.type as m_type, m.id as m_id
        FROM interventions i
        JOIN mannequins m ON i.mannequin_id = m.id
        WHERE i.id = ?
    ''', (intervention_id,)).fetchone()

    if not inter:
        flash('Intervention introuvable.', 'danger')
        conn.close()
        return redirect(url_for('admin_dashboard'))

    photos = conn.execute(
        'SELECT * FROM photos WHERE intervention_id = ? ORDER BY id',
        (intervention_id,)
    ).fetchall()
    conn.close()

    return render_template(
        'admin_intervention.html',
        inter=inter,
        photos=photos
    )


@app.route('/admin/interventions/<int:intervention_id>/delete', methods=['POST'])
@admin_required
def admin_delete_intervention(intervention_id):
    conn = get_db()
    inter = conn.execute(
        'SELECT * FROM interventions WHERE id = ?', (intervention_id,)
    ).fetchone()

    if inter:
        if inter['signature_path']:
            path = os.path.join(SIGNATURES_DIR, inter['signature_path'])
            if os.path.exists(path):
                os.remove(path)
        # Delete associated photos from disk
        inter_photos = conn.execute(
            'SELECT filename FROM photos WHERE intervention_id = ?', (intervention_id,)
        ).fetchall()
        for p in inter_photos:
            path = os.path.join(PHOTOS_DIR, p['filename'])
            if os.path.exists(path):
                os.remove(path)
        conn.execute('DELETE FROM interventions WHERE id = ?', (intervention_id,))
        conn.commit()
        mannequin_id = inter['mannequin_id']
    else:
        mannequin_id = None

    conn.close()
    flash('Intervention supprimée.', 'success')

    if mannequin_id:
        return redirect(url_for('admin_history', mannequin_id=mannequin_id))
    return redirect(url_for('admin_dashboard'))


# ── Excel export ──────────────────────────────────────────

def _format_size(n):
    """Formatte une taille en octets de façon lisible (o / Ko / Mo)."""
    if n < 1024:
        return f"{n} o"
    if n < 1024 * 1024:
        return f"{n / 1024:.0f} Ko"
    return f"{n / (1024 * 1024):.1f} Mo"


def _size_pair(nb, sig_bytes, photo_bytes):
    """(taille sans photos, taille avec photos) formatées, pour nb interventions."""
    overhead = 6 * 1024 + 200 * nb  # structure du .xlsx (approximatif)
    sans = overhead + sig_bytes
    return _format_size(sans), _format_size(sans + photo_bytes)


def _export_sizes():
    """Estime les tailles de l'export Excel (sans/avec photos), au total et par
    mannequin, en un seul passage. Renvoie (paire_totale, {mannequin_id: paire})
    où chaque paire est (taille_sans, taille_avec) déjà formatée."""
    conn = get_db()
    inters = conn.execute(
        "SELECT id, mannequin_id, signature_path FROM interventions"
    ).fetchall()
    photos = conn.execute(
        "SELECT p.filename, i.mannequin_id FROM photos p "
        "JOIN interventions i ON p.intervention_id = i.id"
    ).fetchall()
    conn.close()

    per = {}   # mannequin_id -> [nb, sig_bytes, photo_bytes]
    total = [0, 0, 0]
    for it in inters:
        d = per.setdefault(it['mannequin_id'], [0, 0, 0])
        d[0] += 1
        total[0] += 1
        if it['signature_path']:
            p = os.path.join(SIGNATURES_DIR, it['signature_path'])
            if os.path.exists(p):
                sz = os.path.getsize(p)
                d[1] += sz
                total[1] += sz
    for ph in photos:
        p = os.path.join(PHOTOS_DIR, ph['filename'])
        if os.path.exists(p):
            sz = os.path.getsize(p)
            per.setdefault(ph['mannequin_id'], [0, 0, 0])[2] += sz
            total[2] += sz

    full = _size_pair(*total)
    by_mannequin = {mid: _size_pair(*vals) for mid, vals in per.items()}
    return full, by_mannequin


def _photos_by_intervention(conn, ids):
    """Retourne {intervention_id: [filenames]} pour les interventions données."""
    result = {}
    if not ids:
        return result
    placeholders = ','.join('?' * len(ids))
    rows = conn.execute(
        f'SELECT intervention_id, filename FROM photos '
        f'WHERE intervention_id IN ({placeholders}) ORDER BY id',
        ids
    ).fetchall()
    for r in rows:
        result.setdefault(r['intervention_id'], []).append(r['filename'])
    return result


def _build_tracabilite_xlsx(interventions, sheet_title, doc_title, photos_by_intervention=None):
    """Construit le classeur de traçabilité. Si photos_by_intervention est fourni,
    les photos des interventions sont intégrées dans des colonnes supplémentaires."""
    with_photos = photos_by_intervention is not None
    max_photos = 0
    if with_photos:
        max_photos = max((len(v) for v in photos_by_intervention.values()), default=0)
    last_col = 13 + max_photos  # M = 13 (type de formation), puis une colonne par photo
    last_letter = get_column_letter(last_col)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # nom d'onglet Excel limité à 31 caractères

    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    sub_header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_center = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Ligne 1 : titre
    ws.merge_cells(f'A1:{last_letter}1')
    ws['A1'] = doc_title
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 35

    # Lignes 2-3 : en-têtes
    headers_row2 = {
        'A': 'Date', 'B': 'Prénom', 'C': 'Nom', 'D': 'N° de mannequin',
        'E': 'Nettoyage', 'G': 'Changement des poumons',
        'I': 'Réparation', 'K': 'Informations à communiquer ?', 'L': 'Signature',
        'M': 'Type de formation'
    }
    for col in ['A', 'B', 'C', 'D', 'K', 'L', 'M']:
        ws.merge_cells(f'{col}2:{col}3')
    ws.merge_cells('E2:F2')
    ws.merge_cells('G2:H2')
    ws.merge_cells('I2:J2')

    for col, text in headers_row2.items():
        cell = ws[f'{col}2']
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    sub_headers = {'E': 'oui', 'F': 'non', 'G': 'oui', 'H': 'non', 'I': 'oui', 'J': 'non'}
    for col, text in sub_headers.items():
        cell = ws[f'{col}3']
        cell.value = text
        cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        cell.fill = sub_header_fill
        cell.alignment = center
        cell.border = thin_border

    for row in [2, 3]:
        for col_idx in range(1, 14):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is None and row == 2:
                cell.fill = header_fill
            if row == 3 and col_idx in [1, 2, 3, 4, 11, 12, 13]:
                cell.fill = header_fill
            cell.border = thin_border

    # En-tête "Photos" au-dessus des colonnes photos (à partir de la colonne N)
    if with_photos and max_photos > 0:
        ws.merge_cells(f'N2:{last_letter}3')
        cell = ws['N2']
        cell.value = 'Photos'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        for col_idx in range(14, last_col + 1):
            for r in (2, 3):
                c = ws.cell(row=r, column=col_idx)
                c.fill = header_fill
                c.border = thin_border

    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 20

    col_widths = {'A': 14, 'B': 14, 'C': 14, 'D': 20, 'E': 6, 'F': 6,
                  'G': 6, 'H': 6, 'I': 6, 'J': 6, 'K': 28, 'L': 18, 'M': 22}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    for pj in range(max_photos):
        ws.column_dimensions[get_column_letter(14 + pj)].width = 24

    row_height = 72 if (with_photos and max_photos > 0) else 60

    for idx, inter in enumerate(interventions):
        row = idx + 4
        ws.row_dimensions[row].height = row_height

        data = [
            inter['date'], inter['prenom'], inter['nom'],
            f"{inter['m_type']} N°{inter['m_numero']}",
            'X' if inter['nettoyage'] else '',
            '' if inter['nettoyage'] else 'X',
            'X' if inter['changement_poumons'] else '',
            '' if inter['changement_poumons'] else 'X',
            'X' if inter['reparation'] else '',
            '' if inter['reparation'] else 'X',
            inter['informations'] or '',
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = center if col_idx >= 5 else left_center
            cell.border = thin_border
            cell.font = Font(name='Arial', size=10)

        ws.cell(row=row, column=12).border = thin_border

        # Type de formation (colonne M)
        fcell = ws.cell(row=row, column=13, value=formation_label(inter))
        fcell.alignment = left_center
        fcell.border = thin_border
        fcell.font = Font(name='Arial', size=10)

        # Signature
        if inter['signature_path']:
            sig_file = os.path.join(SIGNATURES_DIR, inter['signature_path'])
            if os.path.exists(sig_file):
                try:
                    img = XlImage(sig_file)
                    img.width = 120
                    img.height = 50
                    ws.add_image(img, f'L{row}')
                except Exception:
                    ws.cell(row=row, column=12, value='[signature]')

        # Photos
        if with_photos:
            for pj, fname in enumerate(photos_by_intervention.get(inter['id'], [])):
                pfile = os.path.join(PHOTOS_DIR, fname)
                if not os.path.exists(pfile):
                    continue
                try:
                    img = XlImage(pfile)
                    target_h = 88
                    if img.height:
                        ratio = target_h / float(img.height)
                        img.height = target_h
                        img.width = int(img.width * ratio)
                    if img.width > 150:
                        ratio2 = 150 / float(img.width)
                        img.width = 150
                        img.height = int(img.height * ratio2)
                    ws.add_image(img, f'{get_column_letter(14 + pj)}{row}')
                except Exception:
                    pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/admin/export')
@admin_required
def admin_export():
    with_photos = request.args.get('photos') == '1'
    conn = get_db()
    interventions = conn.execute('''
        SELECT i.*, m.numero as m_numero, m.type as m_type
        FROM interventions i
        JOIN mannequins m ON i.mannequin_id = m.id
        ORDER BY i.date, m.type, m.numero
    ''').fetchall()
    photos_map = (
        _photos_by_intervention(conn, [i['id'] for i in interventions])
        if with_photos else None
    )
    conn.close()

    output = _build_tracabilite_xlsx(
        interventions, 'Traçabilité',
        'REGISTRE DE TRAÇABILITÉ DES MANNEQUINS', photos_map
    )

    suffix = '_avec_photos' if with_photos else ''
    filename = f"tracabilite_mannequins{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/admin/export/<int:mannequin_id>')
@admin_required
def admin_export_mannequin(mannequin_id):
    """Export Excel for a single mannequin."""
    with_photos = request.args.get('photos') == '1'
    conn = get_db()
    mannequin = conn.execute(
        'SELECT * FROM mannequins WHERE id = ?', (mannequin_id,)
    ).fetchone()

    if not mannequin:
        flash('Mannequin introuvable.', 'danger')
        conn.close()
        return redirect(url_for('admin_dashboard'))

    interventions = conn.execute('''
        SELECT i.*, m.numero as m_numero, m.type as m_type
        FROM interventions i
        JOIN mannequins m ON i.mannequin_id = m.id
        WHERE i.mannequin_id = ?
        ORDER BY i.date
    ''', (mannequin_id,)).fetchall()
    photos_map = (
        _photos_by_intervention(conn, [i['id'] for i in interventions])
        if with_photos else None
    )
    conn.close()

    output = _build_tracabilite_xlsx(
        interventions,
        f"{mannequin['type']} N°{mannequin['numero']}",
        f"TRAÇABILITÉ — {mannequin['type']} N°{mannequin['numero']}",
        photos_map
    )

    suffix = '_avec_photos' if with_photos else ''
    filename = f"tracabilite_{mannequin['type']}_N{mannequin['numero']}{suffix}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ── Signature image serving ──────────────────────────────

@app.route('/signatures/<filename>')
@admin_required
def serve_signature(filename):
    return send_file(os.path.join(SIGNATURES_DIR, filename))


@app.route('/photos/<filename>')
@admin_required
def serve_photo(filename):
    return send_file(os.path.join(PHOTOS_DIR, filename))


# ── Init & run ────────────────────────────────────────────

with app.app_context():
    init_db()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
